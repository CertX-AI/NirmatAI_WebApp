"""Web application UI module of NirmatAI."""

##### IMPORT LIBRARIES #####
import logging
import os
import re
import textwrap
import time
from io import BytesIO
from time import strftime

import mlflow
import pandas as pd
import streamlit as st
from nirmatai_sdk.core import NirmatAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pgpt_python.types import HealthResponse
from streamlit.runtime.uploaded_file_manager import UploadedFile
from utils.css_module import custom_css, local_css
from utils.lock_utils import (
    acquire_lock,
    get_lock_info,
    get_remaining_lock_time,
    is_locked,
    release_lock,
    remove_user_folder,
    update_lock_duration,
)

##### LOGGING SETUP #####
logging.basicConfig(
    level = logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)

logging.getLogger("watchdog").setLevel(logging.WARNING)
logging.getLogger("httpcore.http11").setLevel(logging.WARNING)

logger = logging.getLogger("streamlit")
logger.setLevel(logging.INFO)

logging.info("Streamlit app started")


##### STREAMLIT APP #####
# Set page configuration
st.set_page_config(
    page_title="NirmatAI Submission System",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Create directory for uploaded files
if not os.path.exists("uploaded_files"):
    os.makedirs("uploaded_files")

# Check if the environment variable 'NIRMATAI_BASE_URL' is set
BASE_URL = os.getenv(
    "NIRMATAI_BASE_URL",
    "http://host.containers.internal:8000" # If no URL available
)

logging.info(f"Using BASE_URL: {BASE_URL}")


def print_broken_files(broken_files: list[tuple[str, str]]) -> None:
    """Displays the broken files and their corresponding error messages in Streamlit.

    :param broken_files: A list of tuples, where each tuple contains a file path
                         and an error message.
    """
    if not broken_files:
        st.success("No broken or malformed files encountered in uploaded documents.")
    else:
        broken_files_str = "**The following files encountered issues during ingestion:**\n\n" # noqa: E501
        broken_files_str = broken_files_str + "\n\n".join(
            [f"**File:** {file_path}\n\n**Error:** {error_msg}" for file_path, error_msg in broken_files] # noqa: E501
        )
        st.warning(broken_files_str)


def display_logos(
    sidebar_logo_path: str,
    body_logo_path: str
) -> None:
    """Displays logos in both the sidebar and the main body of a Streamlit app.

    Parameters:
    sidebar_logo_path (str): Logo to be displayed in the sidebar.
    body_logo_path (str): Logo to be displayed in the main body.
    size (str): The size of the logo. Default is 'large'.
    """
    # Display logo in the sidebar and än the main body
    st.logo(sidebar_logo_path, size="large", icon_image=body_logo_path)

# CSS Customization
local_css(custom_css)

# Logo Customization
CERTX_LOGO = "nirmatai_webapp/images/certx_logo.svg"
display_logos(sidebar_logo_path=CERTX_LOGO, body_logo_path=CERTX_LOGO)

# Title of the app
st.title("NirmatAI Submission System")

# Submission status
submission_open = True  # This can be dynamically set based on your backend logic

if submission_open:
    st.success("The system is available to accept submissions.")
else:
    st.error("❌ The system is not available to accept submissions.")

# Initialize session state for login
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
    st.session_state["username"] = ""


# Helper function to convert DataFrame to CSV
@st.cache_data
def convert_df_to_csv(df: pd.DataFrame) -> bytes | None:
    """Convert a DataFrame to a CSV and return it as UTF-8 encoded bytes.

    convert_df caches the result to prevent re-computation on every rerun.
    If the DataFrame is empty or invalid, it returns None.

    Args:
        df (pd.DataFrame): DataFrame to be converted.

    Returns:
        bytes | None: Encoded in UTF-8 as bytes, or None.
    """
    try:
        if df.empty:
            return None
        csv_data = df.to_csv(index=False)
        return csv_data.encode("utf-8") if csv_data else None
    except Exception:
        st.error("Error converting to CSV.")
        return None


# Helper function to convert DataFrame to Excel
def convert_df_to_excel(df: pd.DataFrame) -> bytes | None:
    """Convert a DataFrame to an Excel.

    If the DataFrame is empty or invalid, it returns None.

    Args:
        df (pd.DataFrame): DataFrame to be converted.

    Returns:
        bytes | None: Encoded in UTF-8 as bytes, or None.
    """
    try:
        output = BytesIO()

        # Use 'openpyxl' engine
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Results")

            # Access the workbook and sheet objects
            worksheet = writer.sheets["Results"]

            # Define custom column widths for specific columns
            custom_columns = [
                "Requirement",
                "Potential Means of Compliance",
                "Rationale",
                "Ref. to Doc",
            ]
            default_width = 30
            custom_width = 50

            # Set column widths based on column name
            for idx, col in enumerate(df.columns):
                column_letter = get_column_letter(idx + 1)
                if col in custom_columns:
                    worksheet.column_dimensions[column_letter].width = custom_width
                else:
                    worksheet.column_dimensions[column_letter].width = default_width

            # Set the row height for all rows
            for row in worksheet.iter_rows(
                min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)
            ):
                worksheet.row_dimensions[row[0].row].height = 35

            # Define the columns where the text should be bold
            bold_columns = ["Label", "Compliance status"]
            bold_font = Font(bold=True, size=12, name="Arial")

            # Iterate over each column and apply bold formatting for specific columns
            for idx, col in enumerate(df.columns):
                if col in bold_columns:
                    # Get the column letter
                    column_letter = get_column_letter(idx + 1)
                    # Apply bold font to all cells in this column
                    for cell in worksheet[column_letter]:
                        cell.font = bold_font

            header_font = Font(bold=True, size=14, name="Arial")
            center_alignment = Alignment(horizontal="center")

            for cell in worksheet["1:1"]:  # Apply to the first row (header)
                cell.font = header_font
                cell.alignment = center_alignment

        return output.getvalue()
    except ImportError:
        st.error(
            "The 'openpyxl' module is not installed. Please install it using 'pip install openpyxl'."  # noqa: E501
        )
        return None
    except Exception as e:
        st.error(f"Error converting to Excel: {e!s}")
        return None

def generate_log_file(
    process_logs : list[dict[str, str | int | float | None]],
    file_name : str ="requirement_processing_log.txt"
) -> None:
    """Generate a log file with statistics related to requirement processing.

    :param process_logs: List of dictionaries containing process log data
    :param file_name: Name of the log file to be generated
    """
    # Get the current timestamp
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    # Initialize counters
    total = len(process_logs)
    success_count = sum(
        1 for log in process_logs if log["Processing Status"] == "Success"
    )
    failure_count = total - success_count

    # Calculate total processing time
    total_processing_time = sum(
        log["Processing Time"]
        for log in process_logs
        if isinstance(log["Processing Time"], int | float)  # Ensure numeric type
    )
    hours, remainder = divmod(total_processing_time, 3600)
    minutes, seconds = divmod(remainder, 60)

    # Format total execution time in a verbal format
    formatted_time = f"{int(hours)} hours, {int(minutes)} minutes, {int(seconds)} seconds" if hours > 0 else f"{int(minutes)} minutes, {int(seconds)} seconds" # noqa: E501

    # Calculate average time per requirement
    if total > 0:
        average_time_per_requirement = total_processing_time / total
        avg_minutes, avg_seconds = divmod(average_time_per_requirement, 60)
        formatted_avg_time = f"{int(avg_minutes)} minutes, {int(avg_seconds)} seconds" if avg_minutes > 0 else f"{int(avg_seconds)} seconds" # noqa: E501
    else:
        formatted_avg_time = "No requirements"

    # Format the log
    max_width = 120

    log_content = []
    # Generate the header with dynamic width
    separator = "=" * max_width
    log_content.append(separator)
    log_content.append("NirmatAI Requirement Processing Log File".center(max_width))
    log_content.append(f"Generated on: {timestamp}".center(max_width))
    log_content.append(separator + "\n\n")

    log_content.append("----- Uploaded Documents -----".center(max_width))
    if st.session_state.get("uploaded_docs"):
        # Sort and group uploaded documents by file type
        uploaded_docs = sorted(
            st.session_state["uploaded_docs"],
            key=lambda x: (x.name.split(".")[-1].lower(), x.name.lower())
        )
        grouped_docs : dict[str, list[str]]= {"pdf": [], "docx": [], "txt": []}
        for doc in uploaded_docs:
            ext = doc.name.split(".")[-1].lower()
            if ext in grouped_docs:
                grouped_docs[ext].append(doc.name)

        for file_type, docs in grouped_docs.items():
            if docs:
                log_content.append(f"Uploaded {file_type.upper()} Files:")
                for idx, doc_name in enumerate(docs, start=1):
                    log_content.append(f"   {idx}. Name: {doc_name}")
            else:
                log_content.append(f"No {file_type.upper()} files uploaded.")
    else:
        log_content.append("No uploaded documents.")
    log_content.append("\n")
    log_content.append("----- Uploaded Requirements -----".center(max_width))
    if st.session_state.get("requirements_file_name"):
        log_content.append(f"Name: {st.session_state["requirements_file_name"]}")
    else:
        log_content.append("No uploaded requirements.")

    log_content.append("\n\n")
    log_content.append("----- Summary Statistics -----".center(max_width))
    log_content.append(f"Total Requirements Processed: {total}")
    log_content.append(f"Successfully Processed: {success_count}")
    log_content.append(f"Failed Requirements: {failure_count}")
    log_content.append(f"Processing Duration: {formatted_time}")
    log_content.append(
        f"Average Processing Duration per Requirement: {formatted_avg_time}\n"
    )

    log_content.append("----- Detailed Log -----".center(max_width))
    for idx, log in enumerate(process_logs, start=1):
        # Specify the maximum width for wrapping lines

        log_content.append(f"{idx}. Requirement ID: {log['Requirement ID']}")

        # Format the Requirement field with wrapping
        requirement_text = textwrap.fill(
            str(log["Requirement"]),
            width=max_width - len("  Requirement: "),
            subsequent_indent=" " * 8  # Indent subsequent lines for alignment
        )

        # Append the formatted rationale to the log content
        log_content.append(f"   Requirement: {requirement_text}")
        log_content.append(f"   Compliance Status: {log['Compliance Status']}")

        # Format the Rationale field with wrapping
        rationale_text = textwrap.fill(
            str(log["Rationale"]),
            width=max_width - len("  Rationale: "),
            subsequent_indent=" " * 8  # Indent subsequent lines for alignment
        )

        # Append the formatted rationale to the log content
        log_content.append(f"   Rationale: {rationale_text}")
        log_content.append(f"   Processing Status: {log['Processing Status']}")
        log_content.append(f"   Processing Time: {log['Processing Time']:.2f} seconds")
        log_content.append(f"   Error: {log['Error'] if log['Error'] else 'None'}")
        log_content.append("")

    log_content.append("----- Error Summary -----".center(max_width))
    # Collect all errors in one go
    errors = [
        f"{log['Requirement ID']}: {log['Error'] if log['Error'] else 'No error provided'}" # noqa: E501
        for log in process_logs
        if log["Processing Status"] == "Failure"
    ]

    # Add errors to log_content or a default message if no errors are found
    log_content.extend(errors if errors else ["No error occurred"])


    log_content.append("\n\n" + separator)
    endoffile = "End of the NirmatAI Log File"
    log_content.append(endoffile.center(max_width))
    log_content.append(separator)
    log_content.append(
        "© 2024 NirmatAI CertX AG. All rights reserved.".center(max_width)
    )
    log_content.append(
        "This document contains proprietary information of CertX AG.".center(max_width)
    )
    log_content.append(
        "Unauthorized use, disclosure, or distribution is strictly prohibited.".center(max_width) # noqa: E501
    )

    # Write to file
    with open(file_name, "w") as file:
        file.write("\n".join(log_content))

    logging.info(f"Log file '{file_name}' generated successfully.")

def extract_statistics_from_log(
    log_file: str,
    classes: tuple[int, int, int, int, int]
) -> pd.DataFrame:
    """Extract the statistics section from the log file."""
    statistics = []
    with open(log_file) as file:
        for line in file:
            if "Total Requirements Processed:" in line:
                total_processed = line.split(":")[1].strip()
                statistics.append(
                    ["Total Requirements Processed", int(total_processed)]
                )
            elif "Successfully Processed:" in line:
                success_count = line.split(":")[1].strip()
                statistics.append(["Successfully Processed", int(success_count)])
            elif "Failed Requirements:" in line:
                failed_count = line.split(":")[1].strip()
                statistics.append(["Failed Requirements", int(failed_count)])
            elif "Processing Duration:" in line:
                duration = line.split(":")[1].strip()
                statistics.append(["Processing Duration", duration])
            elif "Average Processing Duration per Requirement:" in line:
                avg_duration = line.split(":")[1].strip()
                statistics.append(["Average Processing Duration", avg_duration])
    statistics.append(["Total Requirements", classes[0]])
    statistics.append(["Full Compliance", classes[1]])
    statistics.append(["Minor Non-Conformity", classes[2]])
    statistics.append(["Major Non-Conformity", classes[3]])
    statistics.append(["Unassigned Compliance", classes[4]])
    return pd.DataFrame(statistics, columns=["General Statistics", "Values"])

def add_sheet_to_excel(
    existing_excel: bytes | None,
    new_sheet_df: pd.DataFrame,
    sheet_name: str
) -> bytes | None:
    """A new sheet to an existing Excel file stored as bytes, with enhanced formatting.

    Args:
        existing_excel (bytes): The existing Excel file in bytes.
        new_sheet_df (pd.DataFrame): The DataFrame to add as a new sheet.
        sheet_name (str): The name of the new sheet to be added.

    Returns:
        bytes: The updated Excel file with the new sheet.
    """
    try:
        # Load the existing Excel workbook from the bytes
        if existing_excel:
            with BytesIO(existing_excel) as input_file:
                workbook = load_workbook(input_file)

            # Check if the sheet name already exists
            if sheet_name in workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' already exists in the workbook."
                )

            # Add the new sheet to the workbook
            sheet = workbook.create_sheet(title=sheet_name)

            # Write DataFrame content to the new sheet
            # Header styling
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4F81BD",
                end_color="4F81BD",
                fill_type="solid"
            )
            center_alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # Apply header styles and write headers
            for col_idx, col_name in enumerate(new_sheet_df.columns, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Write DataFrame rows with alternating colors
            row_fill_1 = PatternFill(
                start_color="D9E2F3",
                end_color="D9E2F3",
                fill_type="solid"
            )
            row_fill_2 = PatternFill(
                start_color="FFFFFF",
                end_color="FFFFFF",
                fill_type="solid"
            )

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row_idx, row in enumerate(new_sheet_df.itertuples(index=False), start=2): # noqa: E501
                fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
                for col_idx, value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = fill
                    cell.border = thin_border

            # Auto-adjust column widths
            for col_idx, col_name in enumerate(new_sheet_df.columns, start=1):
                max_length = max(len(str(col_name)), *(len(str(row[col_idx - 1])) for row in new_sheet_df.itertuples(index=False))) # noqa: E501
                sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2 # noqa: E501

            # Move the new sheet to the first position
            workbook._sheets.insert(0, workbook._sheets.pop(-1))

            # Save the workbook to bytes
            output = BytesIO()
            workbook.save(output)
            output.seek(0)  # Reset pointer for reading
            return output.getvalue()
        else:
            return None

    except Exception as e:
        print(f"Error adding sheet to Excel with formatting: {e}")
        return None


# Helper function to validate file path
def validate_file_path(base_dir: str, filename: str) -> str | None:
    """Validates and returns the full file path if valid; otherwise, None."""
    full_path = os.path.join(base_dir, filename)
    # Ensure the full path is within the base directory and the file exists
    if os.path.commonpath([full_path, base_dir]) == base_dir and os.path.exists(
        full_path
    ):
        return full_path
    else:
        return None


def sanitize_filename(filename: str) -> str:
    """Remove any dangerous characters from the filename.

    That could potentially cause path traversal or other security issues.
    Allow only alphanumeric characters, dashes, and underscores.
    """
    # Replace any character that is an unwanted character
    return re.sub(r"[^A-Za-z0-9_\-.]", "_", filename)


# User authentication (placeholder for real authentication)
def login() -> None:
    """Simulates a user login by setting session state variables.

    This function sets the logged_in state to True and updates
    the username session state with the value from input_username.
    It then triggers a rerun to update the UI based on the new session state.
    """
    st.session_state["logged_in"] = True
    st.session_state["username"] = st.session_state["input_username"]
    st.session_state["current_lockusername"] = None
    st.session_state["NirmatAI_file_name"] = None
    st.session_state["NirmatAI_log_file_name"] = None
    st.session_state["Total_exec_time"]  = None
    st.session_state["Avg_exec_time_per_req"]  = None
    st.rerun()  # Rerun to update the UI


def logout() -> None:
    """Simulates a user logout by resetting session state variables.

    This function sets the logged_in state to False and clears the
    username. It also clears the states for file uploads and resets
    other session states related to document uploads and requirements.
    A rerun is triggered to update the UI based on the new session state.
    """
    remove_user_folder(st.session_state["username"])
    locked = is_locked()
    lock_username, _ = get_lock_info()
    if locked and lock_username == st.session_state["username"]:
        release_lock()
    st.session_state["logged_in"] = False
    st.session_state["username"] = ""
    # Clear file uploader states
    st.session_state["docs_upload"] = None
    st.session_state["req_upload"] = None
    st.session_state["uploaded_docs"] = []
    st.session_state["requirements_df"] = None
    st.session_state["requirements_file_name"] = None
    st.session_state["current_lockusername"] = None
    st.session_state["NirmatAI_file_name"] = None
    st.session_state["NirmatAI_log_file_name"] = None
    st.session_state["Total_exec_time"]  = None
    st.session_state["Avg_exec_time_per_req"]  = None
    st.rerun()  # Rerun to update the UI


# Sidebar content
st.sidebar.title("User Authentication")

if not st.session_state["logged_in"]:
    st.sidebar.text_input(
        "Enter your username",
        key="input_username",
        help="This will be used for submission tracking.",
    )
    if st.sidebar.button("Login"):
        username = st.session_state.get("input_username", "")
        if username:
            # Check if username is at least 8 characters using regex
            if re.fullmatch(r".{8,}", username):
                # Placeholder authentication logic
                logging.info(f"User {username} has logged in to the system")
                login()
            else:
                st.sidebar.error("Username must be at least 8 characters long.")
                st.stop()
        else:
            st.sidebar.error("Please enter your username")
            st.stop()
    else:
        st.warning("You need to log in to submit documents.")
        st.stop()
else:
    st.sidebar.success(f"Logged in as {st.session_state['username']}")
    # Explanation about the system
    st.sidebar.markdown("""
    ### About the System
    - This platform allows you to submit documents and requirements for review.
    - For more information, [CERTX website](https://www.certx.com).
    """)
    if st.sidebar.button("Logout"):
        logging.info(
            f"User {st.session_state['username']} has logged out from the system"
        )
        logout()


def save_uploaded_file(uploadedfile: UploadedFile) -> None:
    """Saves an uploaded file to a user-specific folder.

    This function saves the uploaded file to a folder structure based on
    the username stored in the session state. If the folder does not exist,
    it creates the necessary directories. The file is saved with its original
    filename in binary mode.

    Parameters:
    uploadedfile (UploadedFile): The file object uploaded by the user.
    """
    user_folder = os.path.join(
        "uploaded_files", st.session_state["username"], "documents"
    )
    if not os.path.exists(user_folder):
        os.makedirs(user_folder)
    with open(os.path.join(user_folder, uploadedfile.name), "wb") as f:
        f.write(uploadedfile.getbuffer())


def save_requirements_file(uploadedfile: UploadedFile) -> None:
    """Saves an uploaded requirements file to a user-specific folder.

    This function saves the uploaded file to a folder structure based on
    the username stored in the session state. If the folder does not exist,
    it creates the necessary directories. The file is saved with its original
    filename in binary mode.

    Parameters:
    uploadedfile (UploadedFile): The requirement file object uploaded by the user.
    """
    user_folder = os.path.join(
        "uploaded_files", st.session_state["username"], "requirements"
    )
    results_folder = os.path.join(
        "uploaded_files", st.session_state["username"], "results"
    )
    if not os.path.exists(user_folder):
        os.makedirs(user_folder)
    if not os.path.exists(results_folder):
        os.makedirs(results_folder)
    with open(os.path.join(user_folder, uploadedfile.name), "wb") as f:
        f.write(uploadedfile.getbuffer())


def total_number_of_reqs(file_path: str) -> int:
    """Gets the total number of rows in the single sheet of an Excel file.

    Parameters:
    file_path (str): The path to the Excel file.

    Returns:
    int: The total number of rows in the sheet.
    """
    try:
        # Load the single sheet into a DataFrame
        df = pd.read_excel(file_path)

        # Return the total number of rows
        return len(df)
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return 0


def remove_saved_file(filename: str) -> None:
    """Removes a saved file from the user-specific folder.

    This function deletes the specified file from the user's folder if it exists.

    Parameters:
    filename (str): The name of the file to remove.
    """
    user_folder = os.path.join(
        "uploaded_files", st.session_state["username"], "documents"
    )
    file_path = os.path.join(user_folder, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        # st.info(f"File '{filename}' has been removed.")

def calculate_compliance_statistics(
    results_df: pd.DataFrame
) -> tuple[int, int, int, int, int]:
    """Calculate descriptive statistics for compliance status from a given DataFrame.

    Args:
        results_df (pd.DataFrame): DataFrame containing a 'Compliance status' column.

    Returns:
        tuple[int, int, int, int, int]: A tuple containing:
            - Total number of requirements
            - Number of full compliance entries
            - Number of minor non-conformity entries
            - Number of major non-conformity entries
            - Number of unassigned compliance entries

    Raises:
        ValueError: 'Compliance status' column is missing or input is not a DataFrame.
    """
    try:
        # Validate input type
        if not isinstance(results_df, pd.DataFrame):
            raise TypeError("The input must be a pandas DataFrame.")

        # Check if the required column is present
        if "Compliance status" not in results_df.columns:
            raise ValueError(
                "The DataFrame must contain a 'Compliance status' column."
            )

        # Calculate the statistics
        total_requirements = len(results_df)
        full_compliance = (
            results_df["Compliance status"] == "full-compliance"
        ).sum()
        minor_non_conformity = (
            results_df["Compliance status"] == "minor non-conformity"
        ).sum()
        major_non_conformity = (
            results_df["Compliance status"] == "major non-conformity"
        ).sum()
        unassigned_compliance = results_df["Compliance status"].isna().sum()

        # Log the calculated values for transparency
        logging.info("Calculated statistics successfully:")
        logging.info(f"Total Requirements: {total_requirements}")
        logging.info(f"Full Compliance: {full_compliance}")
        logging.info(f"Minor Non-Conformity: {minor_non_conformity}")
        logging.info(f"Major Non-Conformity: {major_non_conformity}")
        logging.info(f"Unassigned Compliance: {unassigned_compliance}")

        return (
            total_requirements,
            full_compliance,
            minor_non_conformity,
            major_non_conformity,
            unassigned_compliance
        )

    except Exception as e:
        logging.error(
            f"An error occurred while calculating compliance statistics: {e}"
        )
        raise

def process_results() -> None:
    """Processes the results file and offers download options."""
    if st.session_state.get("NirmatAI_file_name") and st.session_state.get("username"):
        try:
            # Construct file path securely
            base_dir = f"uploaded_files/{st.session_state['username']}/results"
            file_path = validate_file_path(
                base_dir, st.session_state["NirmatAI_file_name"]
            )

            if file_path is None:
                st.error("Invalid file path or file does not exist.")
            else:
                # Load the results DataFrame
                results_df = pd.read_csv(file_path)

                # Display header and DataFrame
                st.header("NirmatAI Results for the Processed Requirements")
                st.dataframe(results_df)

                # Add descriptive statistics after displaying the DataFrame
                with st.expander("Descriptive Statistics", expanded=True):
                    # Calculate the statistics
                    compliance_vals = calculate_compliance_statistics(results_df)

                    # Display the statistics using columns and metrics
                    stat_col1, stat_col2, stat_col3, stat_col4, stat_col5 = st.columns(5) # noqa: E501

                    with stat_col1:
                        st.metric(
                            label="Total Requirements",
                            value=compliance_vals[0]
                        )
                    with stat_col2:
                        st.metric(
                            label="Full Compliance",
                            value=compliance_vals[1]
                        )
                    with stat_col3:
                        st.metric(
                            label="Minor Non-conformity",
                            value=compliance_vals[2]
                        )
                    with stat_col4:
                        st.metric(
                            label="Major Non-conformity",
                            value=compliance_vals[3]
                        )
                    with stat_col5:
                        st.metric(
                            label="Unassigned Compliance",
                            value=compliance_vals[4]
                        )
                # Example: Displaying metrics
                if (st.session_state["Total_exec_time"] is not None and
                    st.session_state["Avg_exec_time_per_req"] is not None):
                    with st.expander("Time Statistics", expanded=True):
                        st.metric(
                            label="Total Execution Time",
                            value=st.session_state["Total_exec_time"]
                        )
                        st.metric(
                            label="Average Time per Requirement",
                            value=st.session_state["Avg_exec_time_per_req"]
                        )

                # Convert the DataFrame to CSV, Excel, and HTML formats
                log_file_path = validate_file_path(
                    base_dir, st.session_state["NirmatAI_log_file_name"]
                )
                if log_file_path is None:
                    st.error("Invalid file path or file does not exist.")
                else:
                    statistics_data = extract_statistics_from_log(
                        log_file_path,
                        calculate_compliance_statistics(results_df)
                    )
                    csv_data = convert_df_to_csv(results_df)
                    result_excel = convert_df_to_excel(results_df)
                    # 3. Statistics'i ekle
                    excel_data = add_sheet_to_excel(
                        result_excel,
                        statistics_data,
                        "General Statistics"
                    )
                    html_content = results_df.to_html().encode("utf-8")

                    # Ensure all data is properly converted
                    if csv_data or excel_data or html_content:
                        download_col1, download_col2, download_col3 = st.columns(3)

                        # Download CSV
                        with download_col1:
                            if csv_data is not None:
                                st.download_button(
                                    label="Download the Result as a CSV file",
                                    data=csv_data,
                                    file_name=f"{st.session_state['NirmatAI_file_name']}",
                                    mime="text/csv",
                                )

                        # Download HTML
                        with download_col2:
                            st.download_button(
                                label="Download the Result as an HTML file",
                                data=html_content,
                                file_name=f"NirmatAI_results_{time.strftime('%Y-%m-%d-%H-%M-%S')}.html",
                                mime="text/html",
                            )

                        # Download Excel
                        with download_col3:
                            if excel_data is not None:
                                st.download_button(
                                    label="Download the Result as an Excel file",
                                    data=excel_data,
                                    file_name=f"NirmatAI_results_{time.strftime('%Y-%m-%d-%H-%M-%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                    else:
                        st.error("Failed to convert the results for download.")
        except FileNotFoundError:
            st.error("File not found.")
        except pd.errors.EmptyDataError:
            st.error("The file is empty.")
        except Exception as e:
            st.error(f"An unexpected error occurred: {e}")
    else:
        pass


# Main app tabs
tab1, tab2, tab3 = st.tabs(
    ["📂 Document Upload", "📑 Requirement Upload", "📋 Review & Process"]
)

with tab1:
    st.header("Upload Your Documents")
    st.info("You can upload multiple files at once (PDF, DOCX, TXT).")

    uploaded_docs = st.file_uploader(
        "Choose files",
        type=["pdf", "docx", "txt"],
        accept_multiple_files=True,
        key="docs_upload",
    )

    # Initialize session state for uploaded documents
    if "uploaded_docs" not in st.session_state:
        st.session_state["uploaded_docs"] = []

    # List previously saved files from session state
    previously_uploaded_files = {f.name for f in st.session_state["uploaded_docs"]}

    # If no files are uploaded or files are removed, clear session and remove all files
    if not uploaded_docs and previously_uploaded_files:
        for file_to_remove in previously_uploaded_files:
            remove_saved_file(file_to_remove)
        # Clear the session state for uploaded docs
        st.session_state["uploaded_docs"] = []

    if uploaded_docs:
        current_uploaded_files = {f.name for f in uploaded_docs}

        # Process and save only new uploaded files
        for uploaded_file in uploaded_docs:
            if uploaded_file.name not in previously_uploaded_files:
                save_uploaded_file(uploaded_file)
                # Store file info in session state
                st.session_state["uploaded_docs"].append(uploaded_file)
                logging.info(f"Document {uploaded_file} has been saved to the system")

        # Remove files that were uploaded but no longer in the uploaded_docs
        files_to_remove = previously_uploaded_files - current_uploaded_files
        if files_to_remove:
            for file_to_remove in files_to_remove:
                remove_saved_file(file_to_remove)
                # Update session state to reflect file removal
                st.session_state["uploaded_docs"] = [
                    f
                    for f in st.session_state["uploaded_docs"]
                    if f.name != file_to_remove
                ]

        st.success(
            f"✅ All {len(uploaded_docs)} file(s) have been uploaded successfully."
        )

    # Display list of uploaded files
    if st.session_state["uploaded_docs"]:
        st.subheader("Uploaded Files")
        file_types = {
            "pdf": "📄 PDF Files",
            "docx": "📃 DOCX Files",
            "txt": "📝 TXT Files",
        }
        files_by_type = {}
        for file_type in file_types:
            files_by_type[file_type] = [
                f
                for f in st.session_state["uploaded_docs"]
                if f.name.endswith(file_type)
            ]

        for file_type, files in files_by_type.items():
            st.markdown(f"### {file_types[file_type]}")
            if files:
                for file in files:
                    st.write(f"- {file.name}")
            else:
                st.write("No files uploaded.")
    else:
        st.warning("No documents uploaded yet.")

    # Submit Data button for Tab 1
    if st.button("Submit Documents"):
        if st.session_state["uploaded_docs"]:
            # Placeholder for submission logic
            st.success("🎉 Your documents have been submitted successfully!")
        else:
            st.error("❌ You did not upload any documents.")

with tab2:
    st.header("Upload Your Requirements (Excel files only)")
    st.write("The uploaded Excel file **must** contain the following columns:")
    required_columns = [
        "Requirement Title",
        "Requirement",
        "Requirement ID",
        "Potential Means of Compliance",
    ]
    st.markdown("\n".join([f"- **{col}**" for col in required_columns]))

    # Provide a mockup example to the user for requirements
    st.header("Mockup Example for Requirements")
    mockup_data = {
        "Requirement Title": [
            "Legal responsibility",
            "Certification Agreement",
            "Management of Impartiality",
        ],
        "Requirement": ["Req 1", "Req 2", "Req 3"],
        "Requirement ID": ["5.1.1", "5.1.2", "5.1.3"],
        "Potential Means of Compliance": ["MoC 1", "MoC 2", "MoC 3"],
    }
    mockup_df = pd.DataFrame(mockup_data)
    st.dataframe(mockup_df.style.set_properties(**{"text-align": "left"}))

    st.warning(
        "❗ If these columns are not available, the file upload will be rejected."
    )

    uploaded_requirements = st.file_uploader(
        "Choose an Excel file",
        type=["xlsx"],
        help="Only Excel files (.xlsx) are accepted.",
        key="req_upload",
        accept_multiple_files=False,
    )

    # Initialize session state for requirements
    if "requirements_df" not in st.session_state:
        st.session_state["requirements_df"] = None
        st.session_state["requirements_file_name"] = None
    else:
        # If no file is uploaded or the file is removed
        if not uploaded_requirements:
            st.session_state["requirements_df"] = None
            st.session_state["requirements_file_name"] = None

    requirements_valid = False  # Flag to check if requirements are valid

    if uploaded_requirements is not None:
        try:
            # Load the data
            requirements_df = pd.read_excel(uploaded_requirements)

            # Check if all required columns are in the uploaded file
            if all(col in requirements_df.columns for col in required_columns):
                # st.subheader("🔍 Uploaded Requirements")
                # st.dataframe(requirements_df)
                st.success("✅ Requirements format is correct.")
                requirements_valid = True  # Set flag to True
                # Save the uploaded file
                save_requirements_file(uploaded_requirements)
                if st.session_state["requirements_file_name"] is None:
                    logging.info(
                    f"Requirement file {uploaded_requirements.name} has been saved to the system" # noqa: E501
                )
                # Store requirements in session state
                st.session_state["requirements_df"] = requirements_df
                st.session_state["requirements_file_name"] = uploaded_requirements.name

            else:
                missing_columns = [
                    col
                    for col in required_columns
                    if col not in requirements_df.columns
                ]
                st.error(
                    f"❌ Incorrect format! The following columns are missing: {', '.join(missing_columns)}"  # noqa: E501
                )
                requirements_valid = False  # Set flag to False

        except Exception as e:
            st.error(f"⚠️ Error processing the file: {e}")
            requirements_valid = False  # Set flag to False

    # Display uploaded requirements if available
    if st.session_state["requirements_df"] is not None:
        st.subheader("Uploaded Requirements")
        st.write(f"**File Name:** {st.session_state['requirements_file_name']}")
        st.dataframe(st.session_state["requirements_df"])
    else:
        st.info("No requirements uploaded yet.")

    # Submit Data button for Tab 2
    if st.button("Submit Requirements"):
        if st.session_state["requirements_df"] is not None and requirements_valid:
            # Placeholder for submission logic
            st.success("🎉 Your requirements have been submitted successfully!")
        else:
            st.error("❌ You did not upload any valid requirements.")

with tab3:
    st.header("Review & Process Submissions")

    # Check if the system is locked
    locked = is_locked()
    lock_username, lock_time = get_lock_info()

    if locked:
        if lock_username == st.session_state["username"]:
            # The current user holds the lock
            st.info("You have started processing. Please wait until it completes.")
            logging.info(f"{lock_username} waiting for the process to be finished")
            # Optionally, show progress or status
        else:
            # System is locked by another user
            st.warning(
                f"The system is currently being used by {lock_username}. Please try again later."  # noqa: E501
            )
    else:
        st.info("System is open for Submissions !!!")

    if (
        not st.session_state["uploaded_docs"]
        or st.session_state["requirements_df"] is None
    ):
        st.info(
            "No submissions to display. Please upload documents and requirements first."
        )
    else:
        # Create two columns
        col1, col2 = st.columns(2)

        with col1:
            st.subheader("Uploaded Documents")

            # Check if there are uploaded documents
            if st.session_state["uploaded_docs"]:
                # Use a container for the scrollable section
                st.markdown("", unsafe_allow_html=True)
                st.markdown("", unsafe_allow_html=True)
                st.markdown("", unsafe_allow_html=True)
                with st.container(height=400):
                    for doc in st.session_state["uploaded_docs"]:
                        st.markdown(f"- {doc.name}", unsafe_allow_html=True)
            else:
                st.write("No documents uploaded.")

        with col2:
            st.subheader("Uploaded Requirements")
            if st.session_state["requirements_df"] is not None:
                st.write(f"**File Name:** {st.session_state['requirements_file_name']}")
                st.dataframe(st.session_state["requirements_df"].iloc[:, :5])
            else:
                st.write("No requirements uploaded.")

        # Run the Donwload segment process
        process_results()

        # Process Requirements Button
        if st.session_state["NirmatAI_file_name"] is None and st.button(
            "Process Requirements"
        ):
            # Attempt to acquire the lock atomically
            lock_acquired = acquire_lock()
            if lock_acquired:
                try:
                    if (
                        not st.session_state.get("current_lockusername")
                        or st.session_state.get("current_lockusername")
                        != st.session_state["username"]
                    ):
                        # Set lock ownership in session state
                        st.session_state["current_lockusername"] = st.session_state[
                            "username"
                        ]

                    start_time = strftime("%c")
                    st.success(
                        f"Processing started for: {st.session_state['username']}, Time: {start_time}"  # noqa: E501
                    )
                    time.sleep(2)  # Shorter sleep for better responsiveness
                    logging.info(
                        f"Processing started for: {st.session_state['username']}"
                    )

                    # Validate username input to prevent directory traversal
                    username_safe = sanitize_filename(st.session_state["username"])
                    requirements_safe = sanitize_filename(
                        st.session_state["requirements_file_name"]
                    )

                    logging.info("Filenames are sanitized properly")
                    # Ensure safe experiment name
                    experiment_name = f"User {username_safe} NirmatAI Testing"
                    logging.info("Trying to create the mlflow with set experiment")
                    mlflow.set_experiment(experiment_name)
                    time.sleep(2)  # Shorter sleep for better responsiveness
                    logging.info("Successfully created the mlflow experiment")

                    logging.info("Trying to create the mlflow run")
                    # Simulate processing
                    with (
                        st.spinner("NirmatAI Processing..."),
                        mlflow.start_run(
                            run_name=f"{username_safe} NirmatAI Submission {start_time}"
                        ),
                    ):
                        logging.info("Successfully created the mlflow run")
                        experiment_start_time = time.time()
                        # Initialize the NirmataDemo instance
                        demo = NirmatAI(
                            # system_prompt = syspr,
                            # prompt = uspr,
                            base_url=BASE_URL,
                            timeout=60 * 12,
                            verbose=3,
                        )
                        logging.debug("Successfully created the NirmatAI instance")
                        # Health Check of NirmatAI instance
                        # If status is ko, it means not working
                        if demo.health_check() == HealthResponse(status="ko"):
                            st.warning(
                                "\n\n".join(
                                    [
                                        "❗ Connection Error to NirmatAI: Status Code ==> Not OK",  # noqa: E501
                                        "Please try again later ...",
                                    ]
                                )
                            )
                            logging.error("Connection problem with NirmatAI")
                            time.sleep(15)
                        # If status is NOT ko, it means it is working
                        else:
                            st.info(
                                f"Connection to NirmatAI is successful. Time: {strftime('%c')}"  # noqa: E501
                            )
                            logging.info("Connection to NirmatAI is successful.")

                            # Ensure file path safety by sanitizing paths
                            document_dir = (
                                os.path.join(
                                    "uploaded_files", username_safe, "documents"
                                )
                                + os.path.sep
                            )
                            requirements_path = os.path.join(
                                "uploaded_files",
                                username_safe,
                                "requirements",
                                requirements_safe,
                            )
                            time.sleep(5)
                            try:
                                # Ingest files safely
                                demo.ingest(directory=document_dir)
                                problematic_files = demo.get_broken_files()
                                if not problematic_files:
                                    st.success(
                                        f"✅ Uploaded documents are ingested. Time: {strftime('%c')}"  # noqa: E501
                                    )
                                else:
                                    print_broken_files(problematic_files)
                                    st.success(
                                        f"✅ Other uploaded documents are ingested. Time: {strftime('%c')}"  # noqa: E501
                                    )

                                logging.info("Uploaded documents are ingested.")
                            except Exception as e:
                                st.error(
                                    f"❗ Error occurred while ingesting the documents: {e!s}"  # noqa: E501
                                )
                                logging.error(
                                    f"Error occurred while ingesting the documents: {e!s}" # noqa: E501
                                )

                            # Load requirements and get the results
                            time.sleep(40)
                            demo.load_requirements(reqs_file=requirements_path)
                            mlflow.log_artifact(local_path=requirements_path)

                            # Print number of requirements
                            reqs_number = total_number_of_reqs(requirements_path)
                            st.success(
                                "\n\n".join(
                                    [
                                        "🎉 Your requirement processing has started successfully!",  # noqa: E501
                                        f"Number of Requirements: {reqs_number}",
                                        f"Time: {strftime('%c')}"
                                    ]
                                )
                            )
                            logging.info(
                                "Your requirement has been uploaded successfully."
                            )
                            # Check if lock time needs to be updated
                            locktime_update = update_lock_duration(reqs_number * 5)
                            if locktime_update:
                                logging.info(
                                    f"Lock time has been updated to {reqs_number * 5} minutes" # noqa: E501
                                )

                            try:
                                # Process the requirements
                                logging.info(
                                    "Trying to start requirement processing ..."
                                )
                                result = demo.process_requirements()
                                logging.info(
                                    "Finished requirement processing ..."
                                )
                                st.header(
                                    "NirmatAI Results for the Processed Requirements"
                                )
                                st.dataframe(
                                    result.style.set_properties(
                                        **{"text-align": "left"}
                                    )
                                )

                                # Log the results of the labels
                                c_stats = calculate_compliance_statistics(result)
                                metrics = {
                                    "Total Requirements": c_stats[0],
                                    "Full Compliance": c_stats[1],
                                    "Major Non-Conformity": c_stats[2],
                                    "Minor Non-Conformity": c_stats[3],
                                    "Unassigned Compliance": c_stats[4]
                                }

                                for label, value in metrics.items():
                                    mlflow.log_metric(label, value)
                                    logging.info(f"{label}: {value}")

                                # Save results securely
                                saving_time = time.strftime("%Y-%m-%d-%H:%M:%S")
                                result_file = f"NirmatAI_results_{saving_time}.csv"
                                result_path = os.path.join(
                                    "uploaded_files",
                                    username_safe,
                                    "results",
                                    result_file,
                                )

                                demo.save_results(result, result_path, attach_reqs=True)
                                logging.info(
                                    "Processing Requirements result saved successfully!"
                                )
                                st.success(
                                    f"🎉 Processing Requirements result saved successfully! Time: {strftime('%c')}"  # noqa: E501
                                )
                                st.session_state["NirmatAI_file_name"] = result_file

                                # Saving results as a log
                                result_log_file = f"Result_logs_{saving_time}.txt"
                                result_log_path = os.path.join(
                                    "uploaded_files",
                                    username_safe,
                                    "results",
                                    result_log_file,
                                )
                                result_logs = demo.get_process_logs()
                                st.subheader("Results Overview")
                                st.write(result_logs)
                                generate_log_file(result_logs, str(result_log_path))
                                st.session_state["NirmatAI_log_file_name"] = result_log_file # noqa: E501

                                mlflow.log_artifact(local_path=result_path)
                                logging.info(
                                    "Dataframe logged as an artifact in MLflow."
                                )
                                mlflow.log_artifact(local_path=result_log_path)
                                logging.info(
                                    "Result Logs logged as an artifact in MLflow."
                                )
                            except Exception as e:
                                st.error(
                                    f"❗ Error occurred while processing the requirements: {e!s}"  # noqa: E501
                                )
                                logging.error(
                                    "Error occurred while processing the requirements!"
                                )
                                time.sleep(10)

                            # Delete all ingested documents after use
                            try:
                                demo.delete_all_documents()
                                logging.info("Ingested documents deleted successfully.")
                                st.info("Ingested documents deleted successfully.")
                            except Exception as e:
                                logging.error(
                                    f"An error occurred while deleting the ingested documents: {e!s}"  # noqa: E501
                                )
                                st.error(
                                    f"An error occurred while deleting the ingested documents: {e!s}"  # noqa: E501
                                )

                            # Calculate and display the execution time
                            end_time = time.time()
                            elapsed_time = round(end_time - experiment_start_time, 2)
                            hours, remainder = divmod(elapsed_time, 3600)
                            minutes, seconds = divmod(remainder, 60)

                            # Format total execution time in a verbal format
                            formatted_time = f"{int(hours)} hours, {int(minutes)} minutes, {int(seconds)} seconds" if hours > 0 else f"{int(minutes)} minutes, {int(seconds)} seconds" # noqa: E501

                            if st.session_state["Total_exec_time"] is None:
                                # Store the total execution time in session state
                                st.session_state["Total_exec_time"] = formatted_time

                            # Calculate average time per requirement
                            if reqs_number > 0:
                                average_time_per_requirement = elapsed_time / reqs_number # noqa: E501
                                avg_minutes, avg_seconds = divmod(average_time_per_requirement, 60) # noqa: E501
                                formatted_avg_time = f"{int(avg_minutes)} minutes, {int(avg_seconds)} seconds" if avg_minutes > 0 else f"{int(avg_seconds)} seconds" # noqa: E501
                            else:
                                formatted_avg_time = "No requirements"

                            # Store the average time per requirement in session state
                            if st.session_state["Avg_exec_time_per_req"] is None:
                                st.session_state[
                                    "Avg_exec_time_per_req"
                                ] = formatted_avg_time
                            st.info(
                                f"Execution time: {int(hours)} hours, {int(minutes)} minutes, and {seconds:.2f} seconds"  # noqa: E501
                            )

                            # Log execution time to MLflow
                            mlflow.log_param("Execution Time", elapsed_time)

                            time.sleep(2)
                    logging.info("Processing complete!")
                    st.success("Processing complete!")
                except Exception as e:
                    st.error(f"An unexpected error occurred: {e!s}")
                finally:
                    # Release the lock
                    release_lock()
                    st.rerun()
            else:
                try:
                    # Try to get lock information (who holds it, when it was acquired)
                    lock_username, lock_time = get_lock_info()

                    # Check if the lock username is empty in the session state
                    if not st.session_state.get("current_lockusername") or st.session_state["current_lockusername"] != lock_username: # noqa: E501
                        st.session_state["current_lockusername"] = lock_username
                        st.rerun()
                    else:
                        if lock_time is not None:
                            current_time = time.time()

                            # Ensure the lock time is valid and not in the future
                            if lock_time <= current_time:
                                elapsed_time = current_time - lock_time
                                remaining_time = get_remaining_lock_time()

                                if remaining_time is not None and remaining_time > 0:
                                    # Calculate remaining time
                                    minutes, seconds = divmod(int(remaining_time), 60)
                                    if minutes > 0:
                                        time_message = (
                                            f"{minutes} minutes and {seconds} seconds"
                                        )
                                    else:
                                        time_message = f"{seconds} seconds"

                                    # Display a warning message with remaining time
                                    st.warning(
                                        f"The system is currently being used by {lock_username}. "  # noqa: E501
                                        f"Please try again in {time_message}."
                                    )
                                else:
                                    # Lock expired, system to be unlocked automatically
                                    st.warning(
                                        f"The system was recently locked by {lock_username}, but the lock has expired. "  # noqa: E501
                                        f"Please try again shortly."
                                    )
                                    # Release the lock
                                    release_lock()
                                    time.sleep(2)
                                    st.session_state["current_lockusername"] = None
                                    st.rerun()
                            else:
                                # In case lock_time is somehow in the future
                                st.error(
                                    "An error occurred: lock acquisition time is invalid. Please contact support."  # noqa: E501
                                )
                        else:
                            # lock_time is None (handle gracefully)
                            st.warning(
                                f"The system is currently being used by {lock_username}. "  # noqa: E501
                            )
                except Exception:
                    st.error(
                        "An unexpected error occurred while trying to check the lock status. Please try again later."  # noqa: E501
                    )
                    time.sleep(2)
                    st.rerun()
